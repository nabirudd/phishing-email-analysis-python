# scripts/format_excel.py
# ------------------------------------------------------------
# Beautifies outputs/findings.xlsx with color, filters, width,
# and highlights malicious URLs for presentation/reporting.
# ------------------------------------------------------------

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

IN_FILE = "outputs/findings.xlsx"
OUT_FILE = "outputs/findings_formatted.xlsx"

def format_excel():
    wb = load_workbook(IN_FILE)
    ws = wb.active

    # Header style
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-fit columns and highlight malicious
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
                # Highlight malicious rows
                if cell.row > 1 and cell.value and str(cell.value).isdigit():
                    if int(cell.value) > 0 and ws[f"A{cell.row}"].value:
                        for highlight_cell in ws[cell.row]:
                            highlight_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save as formatted copy
    wb.save(OUT_FILE)
    print(f"✅ Saved beautifully formatted report as {OUT_FILE}")

if __name__ == "__main__":
    format_excel()
